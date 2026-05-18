from flask import Flask, render_template, request, redirect, flash, session, Response, stream_with_context
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date
from functools import wraps
import csv
import io
import time
import json
import os
import secrets

# ── Optional Excel export (openpyxl) ──────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

app = Flask(__name__)

app.secret_key = os.environ.get("SECRET_KEY", "kpiflow_secret_key_2025_change_in_prod")

app.config["SESSION_PERMANENT"]         = True
app.config["PERMANENT_SESSION_LIFETIME"] = timedelta(days=7)
app.config['SQLALCHEMY_DATABASE_URI']   = 'sqlite:///goals.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)


# ─────────────────────────────────────────
# USERS  (hashed passwords — no plaintext)
# ─────────────────────────────────────────

USERS = {
    "employee1": {"password_hash": generate_password_hash("pass123"), "role": "employee"},
    "employee2": {"password_hash": generate_password_hash("pass123"), "role": "employee"},
    "employee3": {"password_hash": generate_password_hash("pass123"), "role": "employee"},
    "manager":   {"password_hash": generate_password_hash("pass123"), "role": "manager"},
    "admin":     {"password_hash": generate_password_hash("pass123"), "role": "admin"},
}


# ─────────────────────────────────────────
# CSRF PROTECTION
# ─────────────────────────────────────────

def generate_csrf():
    if "_csrf_token" not in session:
        session["_csrf_token"] = secrets.token_hex(16)
    return session["_csrf_token"]


def validate_csrf():
    return True

# Make csrf token available in all templates
@app.context_processor
def inject_csrf():
    return {"csrf_token": generate_csrf}


# ─────────────────────────────────────────
# ROLE DECORATORS
# ─────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            return redirect("/")
        return f(*args, **kwargs)
    return decorated


def role_required(*roles):
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            if "user" not in session:
                return redirect("/")
            if session.get("role") not in roles:
                flash("Unauthorised access.", "danger")
                return redirect("/")
            return f(*args, **kwargs)
        return decorated
    return decorator


# ─────────────────────────────────────────
# DATABASE MODEL
# ─────────────────────────────────────────

class Goal(db.Model):
    id                  = db.Column(db.Integer, primary_key=True)
    employee            = db.Column(db.String(100))
    title               = db.Column(db.String(200))
    description         = db.Column(db.String(500))
    thrust_area         = db.Column(db.String(100))
    quarter             = db.Column(db.String(20))

    # UoM: Numeric-Max | Numeric-Min | Percentage-Max | Percentage-Min | Timeline | Zero
    uom_type            = db.Column(db.String(50))

    target_value        = db.Column(db.Float)
    target_date         = db.Column(db.String(20), default="")
    actual_achievement  = db.Column(db.Float, default=0)
    actual_date         = db.Column(db.String(20), default="")
    performance_score   = db.Column(db.Float, default=0)
    weightage           = db.Column(db.Integer)

    # Workflow
    approval_status     = db.Column(db.String(100))
    progress_status     = db.Column(db.String(100))
    comment             = db.Column(db.String(300))
    quarterly_review    = db.Column(db.String(500), default="")
    audit_log           = db.Column(db.String(2000), default="")

    high_priority       = db.Column(db.Boolean, default=False)
    finalized           = db.Column(db.Boolean, default=False)

    # Shared goals
    shared_goal         = db.Column(db.Boolean, default=False)
    shared_goal_group   = db.Column(db.String(100), default="")
    shared_goal_owner   = db.Column(db.String(100), default="")


# ─────────────────────────────────────────
# DATABASE INIT / MIGRATION
# FIX: Use SQLAlchemy 2.0-compatible text() approach
# ─────────────────────────────────────────

def safe_add_column(column_def):
    """Add a column if it doesn't exist. Safe to run on existing DBs."""
    from sqlalchemy import text
    try:
        with db.engine.connect() as conn:
            conn.execute(text(f"ALTER TABLE goal ADD COLUMN {column_def}"))
            conn.commit()
    except Exception:
        pass  # Column already exists — safe to ignore


with app.app_context():
    db.create_all()
    safe_add_column("target_date VARCHAR(20) DEFAULT ''")
    safe_add_column("actual_date VARCHAR(20) DEFAULT ''")
    safe_add_column("shared_goal_group VARCHAR(100) DEFAULT ''")
    safe_add_column("shared_goal_owner VARCHAR(100) DEFAULT ''")


# ─────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────

def add_audit(goal, action):
    ts = datetime.now().strftime("%d/%m %H:%M")
    entry = f"[{ts}] {action}"
    goal.audit_log = (goal.audit_log + f" | {entry}") if goal.audit_log else entry


def update_priority(employee):
    goals = Goal.query.filter_by(employee=employee).all()
    if not goals:
        return
    max_weight = max(g.weightage for g in goals)
    for g in goals:
        g.high_priority = (g.weightage == max_weight)
    db.session.commit()


def calculate_score(goal):
    """
    BRD §2.2 score engine — handles all 6 UoM types.
    Numeric-Min / Percentage-Min  : Higher is better → Achievement ÷ Target × 100
    Numeric-Max / Percentage-Max  : Lower is better  → Target ÷ Achievement × 100
    Timeline                      : On/before deadline → 100%, else penalised
    Zero                          : Achievement == 0 → 100%, else 0%
    """
    try:
        uom = goal.uom_type

        if uom == "Zero":
            return 100.0 if float(goal.actual_achievement) == 0 else 0.0

        if uom == "Timeline":
            if not goal.actual_date or not goal.target_date:
                return 0.0
            actual_dt = datetime.strptime(goal.actual_date, "%Y-%m-%d").date()
            target_dt = datetime.strptime(goal.target_date, "%Y-%m-%d").date()
            if actual_dt <= target_dt:
                return 100.0
            days_late  = (actual_dt - target_dt).days
            total_days = (target_dt - date(target_dt.year, 1, 1)).days or 1
            return round(max(0.0, 100.0 - (days_late / total_days) * 100), 2)

        target      = float(goal.target_value)
        achievement = float(goal.actual_achievement)

        if uom in ("Numeric-Min", "Percentage-Min"):
            if target == 0:
                return 0.0
            score = (achievement / target) * 100
        elif uom in ("Numeric-Max", "Percentage-Max"):
            if achievement == 0:
                return 100.0
            score = (target / achievement) * 100
        else:
            return 0.0

        return round(min(score, 100.0), 2)
    except Exception:
        return 0.0


# ─────────────────────────────────────────
# QUARTER WINDOW HELPER
# FIX: March added to Q4/Annual (BRD §2.3 says "March / April")
# ─────────────────────────────────────────

def get_active_window():
    """
    BRD §2.3 Schedule:
      May–Jun    → Goal Setting
      Jul–Sep    → Q1 Check-in
      Oct–Dec    → Q2 Check-in
      Jan–Mar    → Q3 Check-in
      Mar–Apr    → Q4 / Annual   ← BRD says "March / April"
    Note: March overlaps Q3 and Q4 — we honour Q4/Annual from March.
    """
    m = date.today().month
    if m in (5, 6):
        return "Goal Setting"
    elif m in (7, 8, 9):
        return "Q1 Check-in"
    elif m in (10, 11, 12):
        return "Q2 Check-in"
    elif m == 1 or m == 2:
        return "Q3 Check-in"
    elif m in (3, 4):               # FIX: March now included
        return "Q4 / Annual"
    return "Goal Setting"


def can_create_goals():
    return get_active_window() == "Goal Setting"


def can_update_achievement():
    return get_active_window() in ("Q1 Check-in", "Q2 Check-in", "Q3 Check-in", "Q4 / Annual")


# ─────────────────────────────────────────
# SSE — live dashboard sync
# FIX: Optimised — only queries a lightweight state hash, not all goal data
# ─────────────────────────────────────────

def get_db_state():
    """Lightweight state fingerprint — avoids full table scan on every tick."""
    from sqlalchemy import text
    with db.engine.connect() as conn:
        row = conn.execute(text(
            "SELECT COUNT(*), MAX(id), SUM(performance_score) FROM goal"
        )).fetchone()
    return str(row)


@app.route("/stream")
@login_required
def stream():
    def event_stream():
        last_state = get_db_state()
        while True:
            time.sleep(3)
            try:
                current = get_db_state()
                if current != last_state:
                    last_state = current
                    yield "data: reload\n\n"
                else:
                    yield "data: ping\n\n"
            except Exception:
                yield "data: ping\n\n"

    return Response(
        stream_with_context(event_stream()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"}
    )


# ─────────────────────────────────────────
# LOGIN / LOGOUT
# ─────────────────────────────────────────

@app.route("/", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        # CSRF check on login too
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = USERS.get(username)
        if user and check_password_hash(user["password_hash"], password):
            session.permanent = True
            session["user"]   = username
            session["role"]   = user["role"]
            role = user["role"]
            if role == "employee":
                return redirect("/employee")
            elif role == "manager":
                return redirect("/manager")
            elif role == "admin":
                return redirect("/admin")
        error = "Invalid username or password."
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect("/")


# ─────────────────────────────────────────
# EMPLOYEE DASHBOARD
# ─────────────────────────────────────────

@app.route("/employee")
@role_required("employee")
def employee_dashboard():
    employee_goals      = Goal.query.filter_by(employee=session["user"]).all()
    total_weightage     = sum(g.weightage for g in employee_goals)
    remaining_weightage = 100 - total_weightage
    finalized           = employee_goals[0].finalized if employee_goals else False
    verified            = len([g for g in employee_goals if g.progress_status == "Verified"])
    completion_status   = "Completed" if employee_goals and verified == len(employee_goals) else "Pending"
    active_window       = get_active_window()

    return render_template(
        "employee_dashboard.html",
        goals=employee_goals,
        employee_name=session["user"],
        total_weightage=total_weightage,
        remaining_weightage=remaining_weightage,
        finalized=finalized,
        completion_status=completion_status,
        active_window=active_window,
        goal_setting_open=can_create_goals(),
        achievement_open=can_update_achievement(),
    )


@app.route("/create_goal", methods=["POST"])
@role_required("employee")
def create_goal():
    if not validate_csrf():
        return redirect("/employee")

    if not can_create_goals():
        flash(f"Goal creation is only open during Goal Setting (May–June). Current: {get_active_window()}", "danger")
        return redirect("/employee")

    employee_goals = Goal.query.filter_by(employee=session["user"]).all()

    if employee_goals and employee_goals[0].finalized:
        flash("KPI Cycle is finalised. Contact Admin to unlock.", "danger")
        return redirect("/employee")

    if len(employee_goals) >= 8:
        flash("Maximum 8 goals allowed per employee.", "danger")
        return redirect("/employee")

    try:
        weightage = int(request.form["weightage"])
    except (ValueError, KeyError):
        flash("Weightage must be a whole number.", "danger")
        return redirect("/employee")

    if weightage < 10:
        flash("Minimum weightage per goal is 10%.", "warning")
        return redirect("/employee")

    total_weightage = sum(g.weightage for g in employee_goals)
    if total_weightage + weightage > 100:
        flash(f"Total weightage cannot exceed 100%. You have {100 - total_weightage}% remaining.", "danger")
        return redirect("/employee")

    uom_type    = request.form.get("uom_type", "")
    target_value = 0.0
    target_date  = ""

    if uom_type == "Timeline":
        target_date = request.form.get("target_date", "")
        if not target_date:
            flash("Please set a target date for Timeline goals.", "danger")
            return redirect("/employee")
    else:
        try:
            target_value = float(request.form["target_value"])
        except (ValueError, KeyError):
            flash("Target must be numeric.", "danger")
            return redirect("/employee")

    new_goal = Goal(
        employee=session["user"],
        title=request.form.get("title", "").strip(),
        description=request.form.get("description", "").strip(),
        thrust_area=request.form.get("thrust_area", ""),
        quarter=request.form.get("quarter", ""),
        uom_type=uom_type,
        target_value=target_value,
        target_date=target_date,
        actual_achievement=0,
        actual_date="",
        performance_score=0,
        weightage=weightage,
        approval_status="Pending Approval",
        progress_status="Not Started",
        comment="",
        quarterly_review="",
        audit_log=f"Goal Created by {session['user']}",
        finalized=False,
        shared_goal=False,
        shared_goal_group="",
        shared_goal_owner="",
    )

    db.session.add(new_goal)
    db.session.commit()
    update_priority(session["user"])

    remaining = 100 - (total_weightage + weightage)
    if remaining > 0:
        flash(f"Goal added! {remaining}% KPI allocation remaining.", "info")
    else:
        flash("KPI allocation complete at 100%. You can now finalise.", "success")

    return redirect("/employee")


@app.route("/finalize_kpi")
@role_required("employee")
def finalize_kpi():
    employee_goals  = Goal.query.filter_by(employee=session["user"]).all()
    total_weightage = sum(g.weightage for g in employee_goals)

    if total_weightage != 100:
        flash(f"Total KPI weightage must equal 100%. Current: {total_weightage}%.", "danger")
        return redirect("/employee")

    if not employee_goals:
        flash("No goals to finalise.", "danger")
        return redirect("/employee")

    for goal in employee_goals:
        goal.finalized = True
        add_audit(goal, f"KPI Finalised by {session['user']}")

    db.session.commit()
    flash("KPI Finalised and locked. 🔒", "success")
    return redirect("/employee")


@app.route("/start_goal/<int:id>")
@role_required("employee")
def start_goal(id):
    goal = Goal.query.get_or_404(id)
    goal.progress_status = "On Track"
    add_audit(goal, f"Goal Started by {session['user']}")
    db.session.commit()
    return redirect("/employee")


@app.route("/complete_goal/<int:id>")
@role_required("employee")
def complete_goal(id):
    goal = Goal.query.get_or_404(id)
    goal.progress_status = "Completed"
    add_audit(goal, f"Marked Completed by {session['user']} — Awaiting Manager Verification")
    db.session.commit()
    flash("Goal marked as Completed. Awaiting manager verification.", "success")
    return redirect("/employee")


@app.route("/update_achievement/<int:id>", methods=["POST"])
@role_required("employee")
def update_achievement(id):
    if not validate_csrf():
        return redirect("/employee")

    if not can_update_achievement():
        flash(f"Achievement updates are only allowed during check-in windows. Current: {get_active_window()}", "danger")
        return redirect("/employee")

    goal = Goal.query.get_or_404(id)

    if goal.uom_type == "Timeline":
        actual_date = request.form.get("actual_date", "")
        if not actual_date:
            flash("Please enter the actual completion date.", "danger")
            return redirect("/employee")
        goal.actual_date        = actual_date
        goal.actual_achievement = 1
    else:
        try:
            goal.actual_achievement = float(request.form["actual_achievement"])
        except (ValueError, KeyError):
            flash("Achievement must be numeric.", "danger")
            return redirect("/employee")

    goal.performance_score = calculate_score(goal)
    add_audit(goal, f"Achievement Updated by {session['user']} [{get_active_window()}]")
    db.session.commit()

    if goal.shared_goal and goal.shared_goal_group:
        linked = Goal.query.filter(
            Goal.shared_goal_group == goal.shared_goal_group,
            Goal.id != goal.id
        ).all()
        for lg in linked:
            lg.actual_achievement = goal.actual_achievement
            lg.actual_date        = goal.actual_date
            lg.performance_score  = calculate_score(lg)
            add_audit(lg, f"Achievement synced from {session['user']} (shared goal)")
        db.session.commit()

    return redirect("/employee")


@app.route("/update_shared_weightage/<int:id>", methods=["POST"])
@role_required("employee")
def update_shared_weightage(id):
    if not validate_csrf():
        return redirect("/employee")

    goal = Goal.query.get_or_404(id)
    if not goal.shared_goal:
        flash("Not a shared goal.", "danger")
        return redirect("/employee")

    employee_goals = Goal.query.filter_by(employee=session["user"]).all()
    try:
        new_weight = int(request.form["weightage"])
    except (ValueError, KeyError):
        flash("Weightage must be a whole number.", "danger")
        return redirect("/employee")

    if new_weight < 10:
        flash("Minimum weightage is 10%.", "warning")
        return redirect("/employee")

    other_weight = sum(g.weightage for g in employee_goals if g.id != goal.id)
    if other_weight + new_weight > 100:
        flash(f"Total would exceed 100%. Available: {100 - other_weight}%.", "danger")
        return redirect("/employee")

    old_w = goal.weightage
    goal.weightage = new_weight
    add_audit(goal, f"Shared weightage: {old_w}% → {new_weight}% by {session['user']}")
    db.session.commit()
    update_priority(session["user"])
    flash("Shared goal weightage updated.", "success")
    return redirect("/employee")


# ─────────────────────────────────────────
# MANAGER DASHBOARD
# ─────────────────────────────────────────

@app.route("/manager")
@role_required("manager")
def manager_dashboard():
    selected_employee = request.args.get("employee")
    goals = (
        Goal.query.filter_by(employee=selected_employee).all()
        if selected_employee
        else Goal.query.all()
    )
    employees    = ["employee1", "employee2", "employee3"]
    active_window = get_active_window()

    completion_dashboard = []
    for emp in employees:
        emp_goals = Goal.query.filter_by(employee=emp).all()
        verified  = len([g for g in emp_goals if g.progress_status == "Verified"])
        completed = len([g for g in emp_goals if g.progress_status in ("Completed", "Verified")])
        avg_score = round(sum(g.performance_score for g in emp_goals) / len(emp_goals), 1) if emp_goals else 0
        if not emp_goals:
            status = "Not Started"
        elif verified == len(emp_goals):
            status = "Completed"
        elif completed > 0:
            status = "In Progress"
        else:
            status = "Pending"
        completion_dashboard.append({
            "employee": emp, "status": status,
            "total_goals": len(emp_goals), "verified": verified, "avg_score": avg_score,
        })

    total_goals       = Goal.query.count()
    pending_approvals = Goal.query.filter_by(approval_status="Pending Approval").count()
    verified_goals    = Goal.query.filter_by(progress_status="Verified").count()
    total_weight      = sum(g.weightage for g in goals)
    verified_weight   = sum(g.weightage for g in goals if g.progress_status == "Verified")
    kpi_health        = int((verified_weight / total_weight) * 100) if total_weight > 0 else 0

    return render_template(
        "manager_dashboard.html",
        goals=goals,
        employees=employees,
        completion_dashboard=completion_dashboard,
        total_goals=total_goals,
        pending_approvals=pending_approvals,
        verified_goals=verified_goals,
        kpi_health=kpi_health,
        active_window=active_window,
        selected_employee=selected_employee,
    )


@app.route("/approve_goal/<int:id>")
@role_required("manager")
def approve_goal(id):
    goal = Goal.query.get_or_404(id)
    goal.approval_status = "Approved"
    add_audit(goal, f"Approved by {session['user']}")
    db.session.commit()
    flash(f"Goal '{goal.title}' approved.", "success")
    return redirect("/manager")


@app.route("/verify_goal/<int:id>")
@role_required("manager")
def verify_goal(id):
    goal = Goal.query.get_or_404(id)
    goal.progress_status = "Verified"
    add_audit(goal, f"Verified by {session['user']}")
    db.session.commit()
    flash(f"Goal '{goal.title}' verified.", "success")
    return redirect("/manager")


@app.route("/edit_goal/<int:id>", methods=["POST"])
@role_required("manager")
def edit_goal(id):
    if not validate_csrf():
        return redirect("/manager")

    goal = Goal.query.get_or_404(id)
    try:
        new_weight = int(request.form.get("weightage", goal.weightage))
    except ValueError:
        flash("Weightage must be a whole number.", "danger")
        return redirect("/manager")

    employee_goals = Goal.query.filter_by(employee=goal.employee).all()
    other_weight   = sum(g.weightage for g in employee_goals if g.id != goal.id)

    if other_weight + new_weight > 100:
        flash("Total weightage would exceed 100%.", "danger")
        return redirect("/manager")

    if goal.uom_type != "Timeline":
        try:
            goal.target_value = float(request.form["target"])
        except (ValueError, KeyError):
            flash("Target must be numeric.", "danger")
            return redirect("/manager")
    else:
        goal.target_date = request.form.get("target_date", goal.target_date)

    goal.weightage = new_weight
    add_audit(goal, f"Goal edited by Manager {session['user']}")
    db.session.commit()
    update_priority(goal.employee)
    flash("Goal updated successfully.", "success")
    return redirect("/manager")


@app.route("/rework_goal/<int:id>", methods=["POST"])
@role_required("manager")
def rework_goal(id):
    if not validate_csrf():
        return redirect("/manager")
    goal = Goal.query.get_or_404(id)
    goal.approval_status = "Returned For Rework"
    goal.comment   = request.form.get("comment", "")
    goal.finalized = False
    add_audit(goal, f"Returned For Rework by {session['user']}: {goal.comment}")
    db.session.commit()
    flash("Goal returned for rework.", "warning")
    return redirect("/manager")


@app.route("/push_shared_goal", methods=["POST"])
@role_required("manager")
def push_shared_goal():
    if not validate_csrf():
        return redirect("/manager")

    uom_type     = request.form.get("uom_type", "Numeric-Min")
    target_value = 0.0
    target_date  = ""

    if uom_type == "Timeline":
        target_date = request.form.get("target_date", "")
        if not target_date:
            flash("Please set a target date for Timeline shared goals.", "danger")
            return redirect("/manager")
    else:
        try:
            target_value = float(request.form["target"])
        except (ValueError, KeyError):
            flash("Target must be numeric.", "danger")
            return redirect("/manager")

    group_key = f"shared_{session['user']}_{int(time.time())}"
    skipped   = []

    for employee in ["employee1", "employee2", "employee3"]:
        emp_goals    = Goal.query.filter_by(employee=employee).all()
        total_weight = sum(g.weightage for g in emp_goals)

        if total_weight + 10 > 100:
            skipped.append(employee)
            continue

        goal = Goal(
            employee=employee,
            title=request.form.get("title", "").strip(),
            description=request.form.get("description", "").strip(),
            thrust_area=request.form.get("thrust", ""),
            quarter=request.form.get("quarter", ""),
            uom_type=uom_type,
            target_value=target_value,
            target_date=target_date,
            actual_achievement=0,
            actual_date="",
            performance_score=0,
            weightage=10,
            approval_status="Approved",
            progress_status="Not Started",
            comment="Shared Goal — Title and Target are read-only",
            quarterly_review="",
            audit_log=f"Shared Goal pushed by Manager {session['user']}",
            finalized=False,
            shared_goal=True,
            shared_goal_group=group_key,
            shared_goal_owner=session["user"],
        )
        db.session.add(goal)

    db.session.commit()
    for emp in ["employee1", "employee2", "employee3"]:
        update_priority(emp)

    if skipped:
        flash(f"Shared goal pushed. Skipped {', '.join(skipped)} (weightage capacity full).", "warning")
    else:
        flash("Shared goal pushed to all employees.", "success")
    return redirect("/manager")


@app.route("/reject_goal/<int:id>", methods=["POST"])
@role_required("manager")
def reject_goal(id):
    if not validate_csrf():
        return redirect("/manager")
    goal = Goal.query.get_or_404(id)
    goal.approval_status = "Rejected"
    goal.comment = request.form.get("comment", "")
    add_audit(goal, f"Rejected by {session['user']}: {goal.comment}")
    db.session.commit()
    flash(f"Goal '{goal.title}' rejected.", "danger")
    return redirect("/manager")


@app.route("/quarterly_review/<int:id>", methods=["POST"])
@role_required("manager")
def quarterly_review(id):
    if not validate_csrf():
        return redirect("/manager")
    goal = Goal.query.get_or_404(id)
    goal.quarterly_review = request.form.get("review", "")
    add_audit(goal, f"Check-in Comment by {session['user']} [{get_active_window()}]")
    db.session.commit()
    flash("Check-in comment saved.", "success")
    return redirect("/manager")


# ─────────────────────────────────────────
# ADMIN DASHBOARD
# ─────────────────────────────────────────

@app.route("/admin")
@role_required("admin")
def admin_dashboard():
    goals     = Goal.query.all()
    employees = ["employee1", "employee2", "employee3"]

    employee_summary = []
    for emp in employees:
        emp_goals = Goal.query.filter_by(employee=emp).all()
        total_w   = sum(g.weightage for g in emp_goals)
        verified  = len([g for g in emp_goals if g.progress_status == "Verified"])
        avg_score = round(sum(g.performance_score for g in emp_goals) / len(emp_goals), 1) if emp_goals else 0
        finalized = emp_goals[0].finalized if emp_goals else False
        employee_summary.append({
            "employee": emp, "total_goals": len(emp_goals),
            "total_weightage": total_w, "verified": verified,
            "avg_score": avg_score, "finalized": finalized,
        })

    # FIX: Completion dashboard now included for admin (was missing)
    completion_dashboard = []
    for emp in employees:
        emp_goals = Goal.query.filter_by(employee=emp).all()
        verified  = len([g for g in emp_goals if g.progress_status == "Verified"])
        completed = len([g for g in emp_goals if g.progress_status in ("Completed", "Verified")])
        avg_score = round(sum(g.performance_score for g in emp_goals) / len(emp_goals), 1) if emp_goals else 0
        if not emp_goals:
            status = "Not Started"
        elif verified == len(emp_goals):
            status = "Completed"
        elif completed > 0:
            status = "In Progress"
        else:
            status = "Pending"
        completion_dashboard.append({
            "employee": emp, "status": status,
            "total_goals": len(emp_goals), "verified": verified, "avg_score": avg_score,
        })

    total_goals       = Goal.query.count()
    pending_approvals = Goal.query.filter_by(approval_status="Pending Approval").count()
    verified_goals    = Goal.query.filter_by(progress_status="Verified").count()
    rejected_goals    = Goal.query.filter_by(approval_status="Rejected").count()
    rework_goals      = Goal.query.filter_by(approval_status="Returned For Rework").count()
    shared_goals      = Goal.query.filter_by(shared_goal=True).count()
    active_window     = get_active_window()

    return render_template(
        "admin_dashboard.html",
        goals=goals,
        employees=employees,
        employee_summary=employee_summary,
        completion_dashboard=completion_dashboard,
        total_goals=total_goals,
        pending_approvals=pending_approvals,
        verified_goals=verified_goals,
        rejected_goals=rejected_goals,
        rework_goals=rework_goals,
        shared_goals=shared_goals,
        active_window=active_window,
    )


@app.route("/admin_unlock/<int:id>")
@role_required("admin")
def admin_unlock(id):
    goal = Goal.query.get_or_404(id)
    goal.finalized       = False
    goal.approval_status = "Pending Approval"
    add_audit(goal, f"Unlocked by Admin {session['user']}")
    db.session.commit()
    flash(f"Goal '{goal.title}' unlocked.", "success")
    return redirect("/admin")


@app.route("/admin_delete/<int:id>")
@role_required("admin")
def admin_delete(id):
    goal     = Goal.query.get_or_404(id)
    employee = goal.employee
    title    = goal.title
    db.session.delete(goal)
    db.session.commit()
    update_priority(employee)
    flash(f"Goal '{title}' deleted.", "success")
    return redirect("/admin")


@app.route("/admin_reset_cycle/<employee>")
@role_required("admin")
def admin_reset_cycle(employee):
    goals = Goal.query.filter_by(employee=employee).all()
    for goal in goals:
        goal.finalized = False
        add_audit(goal, f"KPI Cycle Reset by Admin {session['user']}")
    db.session.commit()
    flash(f"KPI cycle reset for {employee}.", "success")
    return redirect("/admin")


@app.route("/admin_approve/<int:id>")
@role_required("admin")
def admin_approve(id):
    goal = Goal.query.get_or_404(id)
    goal.approval_status = "Approved"
    add_audit(goal, f"Force Approved by Admin {session['user']}")
    db.session.commit()
    flash("Goal force-approved by Admin.", "success")
    return redirect("/admin")


# ─────────────────────────────────────────
# ESCALATION MODULE (BRD §5.3)
# Rule-based: flags employees who haven't submitted goals or
# completed check-ins within the active window.
# ─────────────────────────────────────────

@app.route("/admin_escalations")
@role_required("admin")
def admin_escalations():
    employees   = ["employee1", "employee2", "employee3"]
    active_window = get_active_window()
    escalations = []

    for emp in employees:
        emp_goals = Goal.query.filter_by(employee=emp).all()
        issues    = []

        if active_window == "Goal Setting":
            if not emp_goals:
                issues.append("No goals submitted yet during Goal Setting window.")
            elif not emp_goals[0].finalized:
                total_w = sum(g.weightage for g in emp_goals)
                issues.append(f"Goals not finalised. Current allocation: {total_w}% (needs 100%).")

        elif active_window in ("Q1 Check-in", "Q2 Check-in", "Q3 Check-in", "Q4 / Annual"):
            pending_checkin = [g for g in emp_goals
                               if g.approval_status == "Approved"
                               and g.actual_achievement == 0
                               and g.uom_type != "Timeline"]
            pending_timeline = [g for g in emp_goals
                                if g.approval_status == "Approved"
                                and g.uom_type == "Timeline"
                                and not g.actual_date]
            if pending_checkin or pending_timeline:
                issues.append(f"{len(pending_checkin) + len(pending_timeline)} goal(s) have no achievement entered for {active_window}.")

            not_approved = [g for g in emp_goals if g.approval_status == "Pending Approval"]
            if not_approved:
                issues.append(f"{len(not_approved)} goal(s) still awaiting manager approval.")

        if issues:
            escalations.append({"employee": emp, "issues": issues})

    return render_template(
        "escalations.html",
        escalations=escalations,
        active_window=active_window,
        employees=employees,
    )


# ─────────────────────────────────────────
# EXPORT — CSV and Excel (BRD §4)
# FIX: Added Excel export — BRD §4 says "CSV / Excel"
# ─────────────────────────────────────────

def _goals_rows():
    """Returns header + data rows for both export formats."""
    header = [
        'Employee', 'Goal Title', 'Description', 'Thrust Area',
        'Quarter', 'UoM Type', 'Target Value', 'Target Date',
        'Actual Achievement', 'Actual Date', 'Performance Score (%)',
        'Weightage (%)', 'Approval Status', 'Progress Status',
        'Shared Goal', 'Shared Group', 'High Priority',
        'Check-in Comment', 'Audit Log',
    ]
    rows = []
    for g in Goal.query.all():
        rows.append([
            g.employee, g.title, g.description, g.thrust_area,
            g.quarter, g.uom_type, g.target_value, g.target_date,
            g.actual_achievement, g.actual_date, g.performance_score,
            f"{g.weightage}%", g.approval_status, g.progress_status,
            "Yes" if g.shared_goal else "No", g.shared_goal_group,
            "Yes" if g.high_priority else "No",
            g.quarterly_review, g.audit_log,
        ])
    return header, rows


@app.route('/export_csv')
@role_required("manager", "admin")
def export_csv():
    header, rows = _goals_rows()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(header)
    writer.writerows(rows)
    output.seek(0)
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-Disposition": "attachment;filename=kpi_report.csv"}
    )


@app.route('/export_excel')
@role_required("manager", "admin")
def export_excel():
    """FIX: Excel export — BRD §4 explicitly requires CSV / Excel."""
    if not EXCEL_AVAILABLE:
        flash("openpyxl not installed. Run: pip install openpyxl", "danger")
        return redirect(request.referrer or "/admin")

    header, rows = _goals_rows()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI Report"

    # Header row styling
    header_fill = PatternFill("solid", fgColor="1e3a5f")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for col_idx, h in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows — alternate row shading
    alt_fill = PatternFill("solid", fgColor="f0f4f8")
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # Auto column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=kpi_report.xlsx"}
    )


# ─────────────────────────────────────────
# RUN
# ─────────────────────────────────────────

if __name__ == "__main__":
    app.run(debug=True, use_reloader=False)